import json
import re
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtGui import QAction
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QSlider,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QListWidget,
)


APP_TITLE = "Creator OS"
CONFIG_PATH = Path("Settings") / "config.json"
OPENAI_TEST_PROMPT = "Reply only: OK"
DEFAULT_OPENAI_MODEL = "gpt-5.4-mini"
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}
AUDIO_EXTENSIONS = {".mp3"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}


@dataclass
class OpenAIStatus:
    status: str = "Disconnected"
    model: str = DEFAULT_OPENAI_MODEL
    response_time_ms: int | None = None
    requests_today: int = 0
    last_error: str = ""


@dataclass
class ProjectState:
    image_folder: Path | None = None
    image_files: list[Path] = field(default_factory=list)
    audio_folder: Path | None = None
    audio_candidates: dict[str, list[Path]] = field(default_factory=dict)
    selected_audio: dict[str, Path] = field(default_factory=dict)
    lyrics_folder: Path | None = None
    lyrics_by_track: dict[str, str] = field(default_factory=dict)
    preview_duration: int = 0


def load_config() -> dict:
    default_config = {
        "openai": {
            "api_key": "",
            "model": DEFAULT_OPENAI_MODEL,
            "requests_today": 0,
        }
    }
    if not CONFIG_PATH.exists():
        return default_config
    try:
        config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return default_config

    if "openai" not in config:
        old_config = config.get("gemini", {})
        config["openai"] = {
            "api_key": "",
            "model": old_config.get("model", DEFAULT_OPENAI_MODEL) or DEFAULT_OPENAI_MODEL,
            "requests_today": int(old_config.get("requests_today", 0) or 0),
        }
    config["openai"].setdefault("api_key", "")
    config["openai"].setdefault("model", DEFAULT_OPENAI_MODEL)
    config["openai"].setdefault("requests_today", 0)
    return config


def save_config(config: dict) -> None:
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    CONFIG_PATH.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")


def run_openai_connection_test(api_key: str, model: str) -> OpenAIStatus:
    if not api_key.strip():
        return OpenAIStatus(status="Disconnected", model=model, last_error="API key를 입력해 주세요.")

    payload = json.dumps(
        {
            "model": model,
            "input": OPENAI_TEST_PROMPT,
            "max_output_tokens": 16,
        }
    ).encode("utf-8")
    request = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key.strip()}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    started = time.perf_counter()
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            raw = response.read().decode("utf-8", errors="replace")
        elapsed = int((time.perf_counter() - started) * 1000)
        data = json.loads(raw)
        connected_model = data.get("model", model)
        return OpenAIStatus(status="Connected", model=connected_model, response_time_ms=elapsed)
    except urllib.error.HTTPError as error:
        elapsed = int((time.perf_counter() - started) * 1000)
        detail = error.read().decode("utf-8", errors="replace")[:800]
        if error.code == 429:
            return OpenAIStatus(
                status="Rate Limited",
                model=model,
                response_time_ms=elapsed,
                last_error="요청 한도 또는 결제 한도에 도달했습니다.",
            )
        return OpenAIStatus(
            status="Disconnected",
            model=model,
            response_time_ms=elapsed,
            last_error=f"HTTP {error.code}: {detail}",
        )
    except (urllib.error.URLError, TimeoutError, OSError) as error:
        return OpenAIStatus(status="Offline", model=model, last_error=str(error))
    except json.JSONDecodeError as error:
        return OpenAIStatus(status="Disconnected", model=model, last_error=f"응답 파싱 실패: {error}")


def scan_images(folder: Path) -> list[Path]:
    return sorted(path for path in folder.iterdir() if path.suffix.lower() in IMAGE_EXTENSIONS)


def normalize_track(value) -> str:
    text = str(value).strip().lower().replace(" ", "")
    if not text:
        return ""
    digits = "".join(char for char in text if char.isdigit())
    if digits:
        return f"track{int(digits):02d}"
    return text


def track_from_filename(path: Path) -> str:
    match = re.search(r"track[\s_-]*(\d{1,3})", path.stem, re.IGNORECASE)
    if match:
        return f"track{int(match.group(1)):02d}"
    return normalize_track(path.stem)


def scan_audio_candidates(folder: Path) -> dict[str, list[Path]]:
    candidates: dict[str, list[Path]] = {}
    for path in sorted(folder.rglob("*")):
        if path.is_file() and path.suffix.lower() in AUDIO_EXTENSIONS:
            track = track_from_filename(path)
            if track:
                candidates.setdefault(track, []).append(path)
    return dict(sorted(candidates.items()))


def load_lyrics_from_excel_folder(folder: Path) -> dict[str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("openpyxl이 설치되어 있지 않습니다. requirements.txt 설치가 필요합니다.") from error

    lyrics: dict[str, str] = {}
    files = sorted(path for path in folder.iterdir() if path.suffix.lower() in EXCEL_EXTENSIONS)
    for file_path in files:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            track = normalize_track(row[0])
            lyric = "" if row[1] is None else str(row[1]).strip()
            if not track or not lyric or track in {"track", "no", "number"}:
                continue
            lyrics[track] = lyric
        workbook.close()
    return dict(sorted(lyrics.items()))


def tone_colors(tone: str) -> tuple[str, str, str]:
    return {
        "good": ("#065f46", "#d1fae5", "#a7f3d0"),
        "warn": ("#92400e", "#fef3c7", "#fde68a"),
        "bad": ("#991b1b", "#fee2e2", "#fecaca"),
        "neutral": ("#374151", "#f3f4f6", "#e5e7eb"),
    }.get(tone, ("#374151", "#f3f4f6", "#e5e7eb"))


class StatusPill(QLabel):
    def __init__(self, text: str, tone: str = "neutral") -> None:
        super().__init__(text)
        self.setAlignment(Qt.AlignCenter)
        self.setMinimumHeight(28)
        self.set_tone(tone)

    def set_tone(self, tone: str) -> None:
        fg, bg, border = tone_colors(tone)
        self.setStyleSheet(
            f"color: {fg}; background: {bg}; border: 1px solid {border}; "
            "border-radius: 8px; padding: 4px 10px; font-weight: 700;"
        )


class MetricCard(QFrame):
    def __init__(self, title: str, value: str, note: str = "", tone: str = "neutral") -> None:
        super().__init__()
        self.setObjectName("metricCard")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(6)

        title_label = QLabel(title)
        title_label.setObjectName("cardTitle")
        value_label = QLabel(value)
        value_label.setObjectName("cardValue")
        note_label = QLabel(note)
        note_label.setObjectName("cardNote")
        note_label.setWordWrap(True)

        layout.addWidget(title_label)
        layout.addWidget(value_label)
        if note:
            layout.addWidget(note_label)
        layout.addStretch()

        border = {
            "good": "#14b8a6",
            "warn": "#f59e0b",
            "bad": "#ef4444",
            "neutral": "#d1d5db",
        }.get(tone, "#d1d5db")
        self.setStyleSheet(
            f"QFrame#metricCard {{ border: 1px solid {border}; border-radius: 8px; background: white; }}"
        )


class Section(QFrame):
    def __init__(self, title: str, description: str = "") -> None:
        super().__init__()
        self.setObjectName("section")
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(16, 14, 16, 14)
        self.layout.setSpacing(10)

        heading = QLabel(title)
        heading.setObjectName("sectionTitle")
        self.layout.addWidget(heading)
        if description:
            body = QLabel(description)
            body.setObjectName("bodyText")
            body.setWordWrap(True)
            self.layout.addWidget(body)


def openai_tone(status: str) -> str:
    return {
        "Connected": "good",
        "Rate Limited": "warn",
        "Disconnected": "bad",
        "Offline": "bad",
    }.get(status, "neutral")


def button_style(tone: str) -> str:
    if tone == "good":
        return "background: #059669; color: white;"
    if tone == "warn":
        return "background: #d97706; color: white;"
    if tone == "bad":
        return "background: #dc2626; color: white;"
    return ""


class OpenAIDialog(QDialog):
    def __init__(self, config: dict, current_status: OpenAIStatus, parent=None) -> None:
        super().__init__(parent)
        self.config = config
        self.result_status = current_status
        self.setWindowTitle("ChatGPT API 연결 확인")
        self.setMinimumWidth(560)

        openai_config = self.config.setdefault("openai", {})

        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        title = QLabel("ChatGPT API 설정")
        title.setObjectName("dialogTitle")
        layout.addWidget(title)

        info = QLabel("비용 절약 기본값은 gpt-5.4-mini입니다. 더 저렴한 단순 작업은 gpt-5.4-nano도 선택할 수 있습니다.")
        info.setObjectName("bodyText")
        info.setWordWrap(True)
        layout.addWidget(info)

        layout.addWidget(QLabel("API Key"))
        self.api_key = QLineEdit(openai_config.get("api_key", ""))
        self.api_key.setEchoMode(QLineEdit.Password)
        self.api_key.setPlaceholderText("OpenAI API Key 입력")
        layout.addWidget(self.api_key)

        layout.addWidget(QLabel("Model"))
        self.model = QComboBox()
        self.model.addItems(["gpt-5.4-mini", "gpt-5.4-nano", "gpt-5-mini", "gpt-4o-mini"])
        selected_model = openai_config.get("model", current_status.model)
        index = self.model.findText(selected_model)
        if index >= 0:
            self.model.setCurrentIndex(index)
        layout.addWidget(self.model)

        self.status_label = QLabel()
        self.status_label.setObjectName("bodyText")
        layout.addWidget(self.status_label)

        self.error_box = QTextEdit()
        self.error_box.setReadOnly(True)
        self.error_box.setFixedHeight(110)
        layout.addWidget(self.error_box)

        buttons = QHBoxLayout()
        test_button = QPushButton("연결 확인")
        save_button = QPushButton("저장")
        close_button = QPushButton("닫기")
        test_button.clicked.connect(self.test_connection)
        save_button.clicked.connect(self.save_settings)
        close_button.clicked.connect(self.accept)
        buttons.addWidget(test_button)
        buttons.addWidget(save_button)
        buttons.addStretch()
        buttons.addWidget(close_button)
        layout.addLayout(buttons)

        self.update_status(current_status)

    def save_settings(self) -> None:
        openai_config = self.config.setdefault("openai", {})
        openai_config["api_key"] = self.api_key.text().strip()
        openai_config["model"] = self.model.currentText()
        openai_config.setdefault("requests_today", 0)
        save_config(self.config)
        QMessageBox.information(self, "저장 완료", "ChatGPT API 설정을 저장했습니다.")

    def test_connection(self) -> None:
        openai_config = self.config.setdefault("openai", {})
        openai_config["api_key"] = self.api_key.text().strip()
        openai_config["model"] = self.model.currentText()
        save_config(self.config)

        QApplication.setOverrideCursor(Qt.WaitCursor)
        QApplication.processEvents()
        try:
            status = run_openai_connection_test(self.api_key.text(), self.model.currentText())
        finally:
            QApplication.restoreOverrideCursor()

        if status.status in {"Connected", "Rate Limited"}:
            openai_config["requests_today"] = int(openai_config.get("requests_today", 0)) + 1
            status.requests_today = openai_config["requests_today"]
            save_config(self.config)
        else:
            status.requests_today = int(openai_config.get("requests_today", 0))

        self.result_status = status
        self.update_status(status)

    def update_status(self, status: OpenAIStatus) -> None:
        response_time = "-" if status.response_time_ms is None else f"{status.response_time_ms} ms"
        self.status_label.setText(
            f"상태: {status.status}\n"
            f"연결 모델: {status.model}\n"
            f"응답 시간: {response_time}\n"
            f"오늘 연결 확인: {status.requests_today}"
        )
        self.error_box.setPlainText(status.last_error or "Last Error: 없음")


class DashboardPage(QWidget):
    def __init__(self, open_api_dialog) -> None:
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(16)

        header = QHBoxLayout()
        title = QLabel("대시보드")
        title.setObjectName("pageTitle")
        self.openai_pill = StatusPill("ChatGPT: Disconnected", "bad")
        api_button = QPushButton("ChatGPT API")
        api_button.clicked.connect(open_api_dialog)
        header.addWidget(title)
        header.addStretch()
        header.addWidget(self.openai_pill)
        header.addWidget(api_button)
        layout.addLayout(header)

        metrics = QGridLayout()
        metrics.setSpacing(12)
        cards = [
            MetricCard("API 기본 모델", DEFAULT_OPENAI_MODEL, "비용 절약용 mini 기본값", "good"),
            MetricCard("이미지 폴더", "미연결", "영상 이미지 소스"),
            MetricCard("A/B MP3 폴더", "미연결", "TRACK01마다 후보 2개 중 1개 선택"),
            MetricCard("엑셀 가사", "미연결", "왼쪽 track / 오른쪽 가사"),
            MetricCard("Preview", "대기", "페이드 기준 타임라인 확인"),
            MetricCard("Rendering", "대기", "지정 폴더 저장"),
        ]
        for index, card in enumerate(cards):
            metrics.addWidget(card, index // 3, index % 3)
        layout.addLayout(metrics)

        note = Section(
            "작업 순서",
            "이미지 폴더 선택 -> A/B MP3 폴더 선택 -> 트랙별 음원 1개 선택 -> 엑셀 가사 폴더 선택 -> 실행 -> 점검 -> Preview -> Rendering 순서로 진행합니다.",
        )
        layout.addWidget(note, 1)

    def update_openai(self, status: OpenAIStatus) -> None:
        self.openai_pill.setText(f"ChatGPT: {status.status} / {status.model}")
        self.openai_pill.set_tone(openai_tone(status.status))


class VocalStudioPage(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.state = ProjectState()

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(14)

        title = QLabel("영상 제작")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        controls = Section("제작 입력")
        control_row = QHBoxLayout()

        self.image_button = QPushButton("이미지폴더")
        self.image_button.clicked.connect(self.choose_image_folder)
        self.audio_button = QPushButton("A/B MP3 폴더")
        self.audio_button.clicked.connect(self.choose_audio_folder)
        self.lyrics_button = QPushButton("엑셀 가사 폴더")
        self.lyrics_button.clicked.connect(self.choose_lyrics_folder)
        self.run_button = QPushButton("실행")
        self.run_button.clicked.connect(self.run_matching)
        self.check_button = QPushButton("점검")
        self.check_button.clicked.connect(self.run_check)
        self.preview_button = QPushButton("Preview")
        self.preview_button.clicked.connect(self.build_preview)
        self.render_button = QPushButton("Rendering")
        self.render_button.clicked.connect(self.render_project)

        for widget in [
            self.image_button,
            self.audio_button,
            self.lyrics_button,
            self.run_button,
            self.check_button,
            self.preview_button,
            self.render_button,
        ]:
            control_row.addWidget(widget)
        control_row.addStretch()
        controls.layout.addLayout(control_row)
        layout.addWidget(controls)

        status_grid = QGridLayout()
        status_grid.setSpacing(12)
        self.image_status = StatusPill("이미지: 미연결", "bad")
        self.audio_status = StatusPill("A/B 음원: 미연결", "bad")
        self.lyrics_status = StatusPill("가사: 미연결", "bad")
        self.sync_status = StatusPill("싱크: 미점검", "neutral")
        status_grid.addWidget(self.image_status, 0, 0)
        status_grid.addWidget(self.audio_status, 0, 1)
        status_grid.addWidget(self.lyrics_status, 0, 2)
        status_grid.addWidget(self.sync_status, 0, 3)
        layout.addLayout(status_grid)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Track", "선택 MP3", "가사", "이미지"])
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table, 1)

        preview_section = Section("로그 / 미리보기")
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setMinimumHeight(180)
        preview_section.layout.addWidget(self.log)

        slider_row = QHBoxLayout()
        slider_row.addWidget(QLabel("Preview 시간"))
        self.preview_slider = QSlider(Qt.Horizontal)
        self.preview_slider.setRange(0, 0)
        self.preview_slider.valueChanged.connect(self.update_preview_time)
        self.preview_time_label = QLabel("00:00 / 00:00")
        slider_row.addWidget(self.preview_slider, 1)
        slider_row.addWidget(self.preview_time_label)
        preview_section.layout.addLayout(slider_row)
        layout.addWidget(preview_section)

    def append_log(self, message: str) -> None:
        self.log.append(message)

    def choose_image_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "이미지 폴더 선택")
        if not folder:
            return
        self.state.image_folder = Path(folder)
        self.state.image_files = scan_images(self.state.image_folder)
        if self.state.image_files:
            self.image_status.setText(f"이미지: {len(self.state.image_files)}개 연결")
            self.image_status.set_tone("good")
            self.image_button.setStyleSheet(button_style("good"))
            self.append_log(f"[이미지폴더] {self.state.image_folder}")
            self.append_log("[이미지 효과] 첫 화면은 fade-in, 이미지 전환은 fade-out + fade-in으로 처리합니다.")
        else:
            self.image_status.setText("이미지: 파일 없음")
            self.image_status.set_tone("warn")
            self.image_button.setStyleSheet(button_style("warn"))
            self.append_log("[이미지] 선택한 폴더에 이미지 파일이 없습니다.")
        self.refresh_table()

    def choose_audio_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "A/B MP3 폴더 선택")
        if not folder:
            return
        self.state.audio_folder = Path(folder)
        self.state.audio_candidates = scan_audio_candidates(self.state.audio_folder)
        self.state.selected_audio.clear()

        if self.state.audio_candidates:
            total_files = sum(len(files) for files in self.state.audio_candidates.values())
            self.audio_status.setText(f"A/B 음원: {len(self.state.audio_candidates)}트랙 / {total_files}개")
            self.audio_status.set_tone("warn")
            self.audio_button.setStyleSheet(button_style("warn"))
            self.append_log(f"[A/B MP3 폴더] {self.state.audio_folder}")
            self.append_log("[A/B 선택] 각 TRACK 행에서 후보 2개 중 사용할 MP3 1개를 선택해 주세요.")
        else:
            self.audio_status.setText("A/B 음원: MP3 없음")
            self.audio_status.set_tone("bad")
            self.audio_button.setStyleSheet(button_style("bad"))
            self.append_log("[A/B MP3] 선택한 폴더에서 MP3 파일을 찾지 못했습니다.")
        self.refresh_table()

    def choose_lyrics_folder(self) -> None:
        if not self.all_audio_selected():
            QMessageBox.warning(self, "A/B 선택 필요", "가사를 넣기 전에 TRACK별 MP3 후보 중 1개씩 먼저 선택해 주세요.")
            return
        folder = QFileDialog.getExistingDirectory(self, "엑셀 가사 폴더 선택")
        if not folder:
            return
        self.state.lyrics_folder = Path(folder)
        try:
            self.state.lyrics_by_track = load_lyrics_from_excel_folder(self.state.lyrics_folder)
        except RuntimeError as error:
            QMessageBox.critical(self, "가사 로드 실패", str(error))
            return
        if self.state.lyrics_by_track:
            self.lyrics_status.setText(f"가사: {len(self.state.lyrics_by_track)}개 연결")
            self.lyrics_status.set_tone("good")
            self.lyrics_button.setStyleSheet(button_style("good"))
            self.append_log(f"[엑셀 가사] {self.state.lyrics_folder}")
            self.append_log("[형식] 왼쪽 칸 track / 오른쪽 칸 가사 인식 완료")
        else:
            self.lyrics_status.setText("가사: 인식 실패")
            self.lyrics_status.set_tone("warn")
            self.lyrics_button.setStyleSheet(button_style("warn"))
            self.append_log("[엑셀 가사] track/가사 데이터를 찾지 못했습니다.")
        self.refresh_table()

    def refresh_table(self) -> None:
        tracks = sorted(set(self.state.audio_candidates) | set(self.state.lyrics_by_track))
        self.table.setRowCount(len(tracks))
        for row, track in enumerate(tracks):
            self.table.setItem(row, 0, QTableWidgetItem(track))

            combo = QComboBox()
            combo.addItem("선택")
            for candidate in self.state.audio_candidates.get(track, []):
                combo.addItem(candidate.name, str(candidate))
            if track in self.state.selected_audio:
                selected_path = str(self.state.selected_audio[track])
                index = combo.findData(selected_path)
                if index >= 0:
                    combo.setCurrentIndex(index)
            combo.currentIndexChanged.connect(lambda _index, track_name=track, widget=combo: self.set_track_audio(track_name, widget))
            self.table.setCellWidget(row, 1, combo)

            lyric = self.state.lyrics_by_track.get(track, "")
            self.table.setItem(row, 2, QTableWidgetItem(lyric))

            image = self.state.image_files[row % len(self.state.image_files)].name if self.state.image_files else ""
            self.table.setItem(row, 3, QTableWidgetItem(image))

    def set_track_audio(self, track: str, combo: QComboBox) -> None:
        selected = combo.currentData()
        if selected:
            self.state.selected_audio[track] = Path(selected)
        else:
            self.state.selected_audio.pop(track, None)
        self.update_audio_status()

    def all_audio_selected(self) -> bool:
        return bool(self.state.audio_candidates) and all(track in self.state.selected_audio for track in self.state.audio_candidates)

    def update_audio_status(self) -> None:
        total = len(self.state.audio_candidates)
        selected = len(self.state.selected_audio)
        if total and selected == total:
            self.audio_status.setText(f"A/B 음원: {selected}/{total} 선택 완료")
            self.audio_status.set_tone("good")
            self.audio_button.setStyleSheet(button_style("good"))
            self.append_log("[A/B 선택] 모든 트랙의 MP3 선택이 완료되었습니다.")
        elif total:
            self.audio_status.setText(f"A/B 음원: {selected}/{total} 선택")
            self.audio_status.set_tone("warn")
            self.audio_button.setStyleSheet(button_style("warn"))

    def run_matching(self) -> None:
        if not self.state.image_files or not self.state.lyrics_by_track or not self.all_audio_selected():
            QMessageBox.warning(self, "실행 불가", "이미지 폴더, A/B MP3 선택, 엑셀 가사를 순서대로 완료해 주세요.")
            return
        self.append_log("[실행] 선택된 MP3 기준으로 track01부터 가사와 이미지를 매칭했습니다.")
        self.append_log("[전환] 첫 화면 fade-in, 이미지 변경 fade-out + fade-in, 렌더링 계획에는 crossfade로 기록됩니다.")
        self.sync_status.setText("싱크: 실행 완료")
        self.sync_status.set_tone("good")

    def run_check(self) -> None:
        issues: list[str] = []
        if not self.state.image_files:
            issues.append("이미지 파일이 없습니다.")
        if not self.state.audio_candidates:
            issues.append("A/B MP3 파일이 없습니다.")
        if self.state.audio_candidates:
            short_tracks = [track for track, files in self.state.audio_candidates.items() if len(files) < 2]
            if short_tracks:
                issues.append(f"후보가 2개 미만인 트랙: {', '.join(short_tracks)}")
        if not self.all_audio_selected():
            issues.append("트랙별 MP3 선택이 완료되지 않았습니다.")
        if not self.state.lyrics_by_track:
            issues.append("가사 데이터가 없습니다.")
        missing_lyrics = sorted(set(self.state.selected_audio) - set(self.state.lyrics_by_track))
        if missing_lyrics:
            issues.append(f"선택된 음원에 대응하는 가사가 없는 트랙: {', '.join(missing_lyrics)}")
        if self.state.image_files and self.state.lyrics_by_track and len(self.state.image_files) < len(self.state.lyrics_by_track):
            issues.append("이미지 수가 track 수보다 적습니다. 일부 이미지는 반복됩니다.")

        if issues:
            self.sync_status.setText("싱크: 확인 필요")
            self.sync_status.set_tone("warn")
            self.append_log("[점검] 확인 필요")
            for issue in issues:
                self.append_log(f" - {issue}")
        else:
            self.sync_status.setText("싱크: 정상")
            self.sync_status.set_tone("good")
            self.append_log("[점검] 가사, 이미지, 선택 MP3 상태가 정상입니다.")

    def build_preview(self) -> None:
        if not self.state.lyrics_by_track:
            QMessageBox.warning(self, "Preview 불가", "엑셀 가사를 먼저 연결해 주세요.")
            return
        duration_per_track = 30
        self.state.preview_duration = max(duration_per_track, len(self.state.lyrics_by_track) * duration_per_track)
        self.preview_slider.setRange(0, self.state.preview_duration)
        self.preview_slider.setValue(0)
        self.log.clear()
        self.append_log("[Preview] 첫 화면 1.5초 fade-in으로 시작합니다.")
        self.append_log("[Preview] 이미지 전환은 1.0초 fade-out + fade-in으로 자연스럽게 넘어갑니다.")
        for index, (track, lyric) in enumerate(self.state.lyrics_by_track.items()):
            image = self.state.image_files[index % len(self.state.image_files)].name if self.state.image_files else "이미지 없음"
            audio = self.state.selected_audio.get(track)
            audio_name = audio.name if audio else "MP3 미선택"
            start = index * duration_per_track
            end = start + duration_per_track
            effect = "fade-in" if index == 0 else "fade-out + fade-in"
            self.append_log(f"{self.format_time(start)}-{self.format_time(end)} | {track} | {audio_name} | {image} | {effect}")
            self.append_log(f"  {lyric[:120]}")
        self.update_preview_time(0)

    def render_project(self) -> None:
        if not self.state.preview_duration:
            QMessageBox.warning(self, "Rendering 불가", "Preview를 먼저 만들어 주세요.")
            return
        folder = QFileDialog.getExistingDirectory(self, "Rendering 저장 폴더 선택")
        if not folder:
            return
        output_folder = Path(folder)
        tracks = list(self.state.lyrics_by_track.items())
        plan = {
            "image_folder": str(self.state.image_folder) if self.state.image_folder else "",
            "audio_folder": str(self.state.audio_folder) if self.state.audio_folder else "",
            "lyrics_folder": str(self.state.lyrics_folder) if self.state.lyrics_folder else "",
            "duration_seconds": self.state.preview_duration,
            "opening_effect": {"type": "fade_in", "duration_seconds": 1.5},
            "image_transition": {"type": "fade_out_fade_in", "duration_seconds": 1.0},
            "tracks": [
                {
                    "track": track,
                    "lyric": lyric,
                    "audio": str(self.state.selected_audio.get(track, "")),
                    "image": str(self.state.image_files[index % len(self.state.image_files)])
                    if self.state.image_files
                    else "",
                    "start_seconds": index * 30,
                    "duration_seconds": 30,
                }
                for index, (track, lyric) in enumerate(tracks)
            ],
        }
        target = output_folder / "render_plan.json"
        target.write_text(json.dumps(plan, indent=2, ensure_ascii=False), encoding="utf-8")
        self.append_log(f"[Rendering] 저장 완료: {target}")
        QMessageBox.information(self, "Rendering 완료", f"렌더링 계획을 저장했습니다.\n{target}")

    def update_preview_time(self, value: int) -> None:
        self.preview_time_label.setText(f"{self.format_time(value)} / {self.format_time(self.state.preview_duration)}")

    @staticmethod
    def format_time(seconds: int) -> str:
        minutes, secs = divmod(max(0, seconds), 60)
        return f"{minutes:02d}:{secs:02d}"


class PlaceholderPage(QWidget):
    def __init__(self, title: str, description: str) -> None:
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(16)
        heading = QLabel(title)
        heading.setObjectName("pageTitle")
        layout.addWidget(heading)
        section = Section(title, description)
        layout.addWidget(section, 1)


class SettingsPage(QWidget):
    def __init__(self, open_api_dialog) -> None:
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(16)
        heading = QLabel("설정")
        heading.setObjectName("pageTitle")
        layout.addWidget(heading)

        openai = Section("ChatGPT API", "API key와 모델을 입력하고 연결된 모델명을 확인합니다.")
        openai_button = QPushButton("ChatGPT API 입력 및 연결 확인")
        openai_button.setMinimumHeight(40)
        openai_button.clicked.connect(open_api_dialog)
        openai.layout.addWidget(openai_button)
        layout.addWidget(openai)
        layout.addStretch()


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.config = load_config()
        openai_config = self.config.setdefault("openai", {})
        self.openai_status = OpenAIStatus(
            model=openai_config.get("model", DEFAULT_OPENAI_MODEL),
            requests_today=int(openai_config.get("requests_today", 0)),
        )

        self.setWindowTitle(APP_TITLE)
        self.resize(1240, 820)

        root = QWidget()
        root_layout = QHBoxLayout(root)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        sidebar = QWidget()
        sidebar.setObjectName("sidebar")
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(14, 16, 14, 14)
        sidebar_layout.setSpacing(12)
        sidebar.setFixedWidth(240)

        brand = QLabel("Creator OS")
        brand.setObjectName("brand")
        subtitle = QLabel("AI 영상 제작 워크스페이스")
        subtitle.setObjectName("sidebarSub")
        sidebar_layout.addWidget(brand)
        sidebar_layout.addWidget(subtitle)

        self.nav = QListWidget()
        self.nav.setObjectName("nav")
        sidebar_layout.addWidget(self.nav, 1)

        self.stack = QStackedWidget()
        self.dashboard_page = DashboardPage(self.open_openai_dialog)
        pages = [
            ("대시보드", self.dashboard_page),
            ("영상 제작", VocalStudioPage()),
            ("휴식 음악", PlaceholderPage("휴식 음악", "Ambient / Journey 영상 제작 흐름은 다음 단계에서 연결합니다.")),
            ("쇼츠 제작", PlaceholderPage("쇼츠 제작", "Hook Finder와 세로 영상 제작 흐름은 다음 단계에서 연결합니다.")),
            ("라이브러리", PlaceholderPage("라이브러리", "이미지, 가사, 음원, 영상 자산을 모아 관리합니다.")),
            ("분석", PlaceholderPage("분석", "성과 데이터와 제작 메모리를 기록합니다.")),
            ("설정", SettingsPage(self.open_openai_dialog)),
        ]
        for name, page in pages:
            self.nav.addItem(name)
            self.stack.addWidget(page)

        self.nav.currentRowChanged.connect(self.stack.setCurrentIndex)
        self.nav.setCurrentRow(1)

        root_layout.addWidget(sidebar)
        root_layout.addWidget(self.stack, 1)
        self.setCentralWidget(root)

        api_action = QAction("ChatGPT API 연결 확인", self)
        api_action.triggered.connect(self.open_openai_dialog)
        exit_action = QAction("종료", self)
        exit_action.triggered.connect(self.close)
        file_menu = self.menuBar().addMenu("파일")
        file_menu.addAction(api_action)
        file_menu.addSeparator()
        file_menu.addAction(exit_action)

        self.statusBar().showMessage("Creator OS 준비 완료")
        self.dashboard_page.update_openai(self.openai_status)
        self.setStyleSheet(APP_STYLES)

    def open_openai_dialog(self) -> None:
        dialog = OpenAIDialog(self.config, self.openai_status, self)
        dialog.exec()
        self.openai_status = dialog.result_status
        self.dashboard_page.update_openai(self.openai_status)
        self.statusBar().showMessage(f"ChatGPT 상태: {self.openai_status.status} / {self.openai_status.model}")


APP_STYLES = """
QMainWindow { background: #f5f7fb; }
QWidget#sidebar {
    background: #101828;
    color: #e5e7eb;
}
QLabel#brand {
    color: white;
    font-size: 22px;
    font-weight: 800;
}
QLabel#sidebarSub {
    color: #9ca3af;
    font-size: 12px;
}
QListWidget#nav {
    background: transparent;
    color: #d1d5db;
    border: none;
    font-size: 14px;
    outline: none;
}
QListWidget#nav::item {
    min-height: 40px;
    padding: 8px 10px;
    border-radius: 6px;
}
QListWidget#nav::item:selected {
    background: #2563eb;
    color: white;
}
QListWidget#nav::item:hover {
    background: #1f2937;
}
QLabel#pageTitle {
    font-size: 28px;
    font-weight: 800;
    color: #111827;
}
QLabel#dialogTitle {
    font-size: 22px;
    font-weight: 800;
    color: #111827;
}
QLabel#sectionTitle {
    color: #111827;
    font-size: 16px;
    font-weight: 800;
}
QLabel#bodyText {
    font-size: 13px;
    color: #4b5563;
}
QLabel#cardTitle {
    color: #6b7280;
    font-size: 12px;
    font-weight: 700;
}
QLabel#cardValue {
    color: #111827;
    font-size: 24px;
    font-weight: 800;
}
QLabel#cardNote {
    color: #6b7280;
    font-size: 12px;
}
QFrame#section {
    background: white;
    border: 1px solid #dbe1ea;
    border-radius: 8px;
}
QPushButton {
    background: #2563eb;
    color: white;
    border: none;
    border-radius: 6px;
    padding: 8px 12px;
    font-weight: 700;
}
QPushButton:hover { background: #1d4ed8; }
QLineEdit, QComboBox, QTextEdit, QTableWidget {
    background: white;
    border: 1px solid #d1d5db;
    border-radius: 6px;
    padding: 8px;
    color: #111827;
}
QHeaderView::section {
    background: #eef2ff;
    color: #111827;
    border: none;
    padding: 8px;
    font-weight: 700;
}
QMenuBar {
    background: #ffffff;
    color: #111827;
}
QStatusBar {
    background: #ffffff;
    color: #374151;
}
"""


def main() -> None:
    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
